"""
Excel Exporter
==============

Exports data to Excel (.xlsx) with formatting.
"""

from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

from loguru import logger

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    logger.warning("openpyxl not installed. Excel export will not be available.")


class ExcelExporter:
    """
    Exports data to Excel (.xlsx) format with professional styling.
    
    Features:
    - Auto-sized columns
    - Styled header row
    - Number formatting
    - Date formatting
    - Filtering enabled
    """
    
    def __init__(self, config: Optional[Dict[str, Any]] = None):
        """
        Initialize the Excel exporter.
        
        Args:
            config: Export configuration options.
        """
        self.config = config or {}
        self.sheet_name = self.config.get("sheet_name", "Data")
        self.include_metadata = self.config.get("include_metadata", True)
        
        # Styles
        self.header_font = Font(bold=True, color="FFFFFF", size=11)
        self.header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        self.cell_alignment = Alignment(vertical="center")
        self.number_format = '#,##0.00'
        self.date_format = 'YYYY-MM-DD'
        
        self.border = Border(
            left=Side(style='thin', color='D9D9D9'),
            right=Side(style='thin', color='D9D9D9'),
            top=Side(style='thin', color='D9D9D9'),
            bottom=Side(style='thin', color='D9D9D9')
        )
    
    def export(
        self,
        data: List[Dict[str, Any]],
        output_dir: str,
        filename: str = "output"
    ) -> Path:
        """
        Export data to Excel file.
        
        Args:
            data: List of dictionaries to export.
            output_dir: Directory to save the file.
            filename: Base filename (without extension).
            
        Returns:
            Path to the created Excel file.
        """
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export. Install with: pip install openpyxl")
        
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        output_file = output_path / f"{filename}.xlsx"
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = self.sheet_name
        
        if not data:
            wb.save(output_file)
            logger.info(f"Created empty Excel file: {output_file}")
            return output_file
        
        # Get headers (exclude internal fields starting with _)
        headers = self._get_headers(data)
        
        # Write headers
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data rows
        for row_idx, row_data in enumerate(data, start=2):
            for col_idx, header in enumerate(headers, start=1):
                value = row_data.get(header, "")
                
                # Clean and format value
                formatted_value = self._format_value(value)
                
                cell = ws.cell(row=row_idx, column=col_idx, value=formatted_value)
                cell.alignment = self.cell_alignment
                cell.border = self.border
                
                # Apply number format if numeric
                if isinstance(formatted_value, (int, float)):
                    cell.number_format = self.number_format
                elif isinstance(formatted_value, datetime):
                    cell.number_format = self.date_format
        
        # Auto-size columns
        self._autosize_columns(ws, headers)
        
        # Add auto-filter
        if len(data) > 0:
            ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data) + 1}"
        
        # Freeze header row
        ws.freeze_panes = "A2"
        
        # Add metadata sheet if configured
        if self.include_metadata:
            self._add_metadata_sheet(wb, data, filename)
        
        # Save workbook
        wb.save(output_file)
        
        logger.info(f"Excel exported: {output_file} ({len(data)} rows)")
        
        return output_file
    
    def _get_headers(self, data: List[Dict[str, Any]]) -> List[str]:
        """Get all unique headers from data, excluding internal fields."""
        headers = []
        seen = set()
        
        for row in data:
            for key in row.keys():
                if key not in seen and not key.startswith('_'):
                    headers.append(key)
                    seen.add(key)
        
        return headers
    
    def _format_value(self, value: Any) -> Any:
        """Format a value for Excel cell."""
        if value is None:
            return ""
        
        if isinstance(value, bool):
            return "Yes" if value else "No"
        
        if isinstance(value, (list, dict)):
            return str(value)
        
        if isinstance(value, str):
            # Try to detect and convert dates
            if self._looks_like_date(value):
                try:
                    from dateutil import parser
                    return parser.parse(value)
                except:
                    pass
            
            # Try to convert to number
            if self._looks_like_number(value):
                try:
                    # Remove currency symbols and separators
                    clean = value.replace('$', '').replace(',', '').replace(' ', '')
                    return float(clean)
                except:
                    pass
        
        return value
    
    def _looks_like_date(self, value: str) -> bool:
        """Check if value looks like a date."""
        import re
        date_patterns = [
            r'^\d{4}-\d{2}-\d{2}',  # ISO format
            r'^\d{2}/\d{2}/\d{4}',  # DD/MM/YYYY
        ]
        return any(re.match(p, value) for p in date_patterns)
    
    def _looks_like_number(self, value: str) -> bool:
        """Check if value looks like a number."""
        import re
        # Match numbers with optional currency, thousands separator
        pattern = r'^[\$\-]?[\d,\.]+$'
        clean = value.replace(' ', '')
        return bool(re.match(pattern, clean))
    
    def _autosize_columns(self, ws, headers: List[str]) -> None:
        """Auto-size columns based on content."""
        for col_idx, header in enumerate(headers, start=1):
            col_letter = get_column_letter(col_idx)
            
            # Start with header width
            max_width = len(str(header)) + 2
            
            # Check first 100 rows for max width
            for row in range(2, min(102, ws.max_row + 1)):
                cell_value = ws.cell(row=row, column=col_idx).value
                if cell_value:
                    cell_width = len(str(cell_value))
                    max_width = max(max_width, min(cell_width + 2, 50))  # Cap at 50
            
            ws.column_dimensions[col_letter].width = max_width
    
    def _add_metadata_sheet(
        self,
        wb: 'Workbook',
        data: List[Dict[str, Any]],
        filename: str
    ) -> None:
        """Add a metadata sheet with export information."""
        ws = wb.create_sheet("Metadata")
        
        metadata = [
            ("Export Information", ""),
            ("", ""),
            ("Filename", filename),
            ("Export Date", datetime.now().strftime("%Y-%m-%d %H:%M:%S")),
            ("Total Rows", len(data)),
            ("Total Columns", len(self._get_headers(data))),
            ("", ""),
            ("Generated by", "PDF to Spreadsheet Automation"),
        ]
        
        for row_idx, (label, value) in enumerate(metadata, start=1):
            ws.cell(row=row_idx, column=1, value=label)
            ws.cell(row=row_idx, column=2, value=value)
            
            if row_idx == 1:
                ws.cell(row=1, column=1).font = Font(bold=True, size=14)
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 40
